from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import load_workbook

app = Flask(__name__)
CORS(app)  # This will enable CORS for all routes

@app.route('/receive_data', methods=['POST'])
def receive_data():
    if request.method == 'POST':
        try:
            # Load the Excel file
            wb = load_workbook('DOC-20240425-WA0016..xlsx')
            ws = wb.active
            
            # Get the values from the received JSON data
            data_values = list(request.json.values())
            print(data_values)
            
            # Loop through each row in the Excel sheet
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Convert row values to strings
                row_values = [str(cell) for cell in row[:-2]]  # Exclude the last three columns
                #print(row_values)
                label = row_values[-1]
                row_values=row_values[:7]
                print(row_values)
               ## print(label)
                  
                # Compare the row values with the received JSON data values
                if row_values == data_values:
                    print("inside")
                    response_data = {
                        "response": "Data received successfully",
                        "label": label
                    }
                    return jsonify(response_data)  # Return the response immediately upon finding a match
                   
            # If no match is found after checking all rows, return a response indicating no matching data
            response_data = {
                "response": "No matching data found",
                "label": None
            }
            return jsonify(response_data)
        
        except Exception as e:
            return jsonify({"error": str(e)})

if __name__ == '__main__':
    app.run(debug=True, port=12345)
